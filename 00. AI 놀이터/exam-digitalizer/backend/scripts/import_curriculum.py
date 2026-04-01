"""교과과정 표준체계 + 학습맵 임포트 스크립트

실행: docker-compose exec backend python scripts/import_curriculum.py

소스:
  - 교과과정_표준체계_AI용 (1).xlsx → curriculum_standards (420행)
  - 학습맵__AI용.xlsx → learning_maps (1,821행) + learning_map_standards (N:M)
"""
import asyncio
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import structlog
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.orm import sessionmaker

from config import settings
from models.curriculum import CurriculumStandard, LearningMap, LearningMapStandard

logger = structlog.get_logger()

# xlsx 파일 경로 (Docker 내에서 접근 가능해야 함)
CURRICULUM_XLSX = os.environ.get(
    "CURRICULUM_XLSX",
    "/data/교과과정_표준체계_AI용 (1).xlsx",
)
LEARNING_MAP_XLSX = os.environ.get(
    "LEARNING_MAP_XLSX",
    "/data/학습맵__AI용.xlsx",
)


def _load_xlsx(path: str, sheet_name: str, header_row: int = 1) -> list[dict]:
    """xlsx 파일을 dict 목록으로 로드"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    headers = []
    for cell in ws[header_row]:
        headers.append(str(cell.value) if cell.value else f"col{len(headers)}")

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        data = {}
        for i, val in enumerate(row):
            if i < len(headers) and val is not None:
                data[headers[i]] = val
        if data:
            rows.append(data)
    return rows


async def import_curriculum_standards(db: AsyncSession, xlsx_path: str) -> dict[str, int]:
    """표준체계 임포트 → {standard_id: pk} 매핑 반환"""
    rows = _load_xlsx(xlsx_path, "교육과정표준체계")
    id_map: dict[str, int] = {}
    count = 0

    for row in rows:
        std_id = str(row.get("교육과정표준체계ID", "")).strip()
        if not std_id:
            continue

        # 중복 체크
        existing = await db.execute(
            select(CurriculumStandard).where(CurriculumStandard.standard_id == std_id)
        )
        if existing.scalar_one_or_none():
            result = await db.execute(
                select(CurriculumStandard.id).where(CurriculumStandard.standard_id == std_id)
            )
            id_map[std_id] = result.scalar_one()
            continue

        cs = CurriculumStandard(
            standard_id=std_id,
            revision_year=int(row.get("개정년도", 2022)),
            subject_group=str(row.get("교과(군)", "")),
            subject=str(row.get("과목", "수학")),
            grade_group=str(row.get("학년(군)", "")),
            content_area=str(row.get("내용체계 영역", "")),
            content_element_1=str(row.get("1단계 내용요소", "")) or None,
            content_element_2=str(row.get("2단계 내용요소", "")) or None,
            content_element_3=str(row.get("3단계 내용요소", "")) or None,
            achievement_code=str(row.get("성취기준 코드", "")) or None,
            achievement_desc=str(row.get("성취기준 내용", "")) or None,
            status=str(row.get("상태", "사용")),
            knowledge_map_ids=str(row.get("지식맵ID", "")) or None,
        )
        db.add(cs)
        count += 1

    await db.flush()

    # PK 매핑 조회
    result = await db.execute(select(CurriculumStandard))
    for cs in result.scalars():
        id_map[cs.standard_id] = cs.id

    logger.info("curriculum_standards_imported", count=count, total=len(id_map))
    return id_map


async def import_learning_maps(
    db: AsyncSession, xlsx_path: str, std_id_map: dict[str, int],
) -> None:
    """학습맵 임포트 + 표준체계 N:M 매핑"""
    rows = _load_xlsx(xlsx_path, "학습맵(수학)", header_row=2)
    lm_count = 0
    link_count = 0

    for row in rows:
        lm_id = str(row.get("학습맵ID", "")).strip()
        if not lm_id:
            continue

        # 중복 체크
        existing = await db.execute(
            select(LearningMap).where(LearningMap.learning_map_id == lm_id)
        )
        if existing.scalar_one_or_none():
            continue

        # Depth 숫자들
        d1 = str(row.get("Depth1_number", "00"))
        d2 = str(row.get("Depth2_number", "")) or None
        d3 = str(row.get("Depth3_number", "")) or None
        d4 = str(row.get("Depth4_number", "")) or None

        # is_leaf: Depth3 또는 Depth4가 있고 이름이 있으면 leaf
        is_leaf = bool(
            (d3 and d3 != "00" and row.get("Depth3_단원명"))
            or (d4 and d4 != "00" and row.get("Depth4_단원명"))
        )

        lm = LearningMap(
            learning_map_id=lm_id,
            short_id=str(row.get("col0", "")) or None,
            revision_year=int(row.get("개정연도", 2022)),
            publish_year=int(row.get("발행연도", 2025)),
            subject_group=str(row.get("과목군", "MAT")),
            subject_code=str(row.get("과목코드", "MAT")),
            school_level=str(row.get("교급", "E")),
            grade=int(row.get("학년", 0)),
            semester=int(row.get("학기", 0)),
            map_number=str(row.get("학습맵 번호", "")),
            depth1_number=d1,
            depth1_name=str(row.get("Depth1_단원명", "")) or None,
            depth2_number=d2,
            depth2_name=str(row.get("Depth2_단원명", "")) or None,
            depth3_number=d3,
            depth3_name=str(row.get("Depth3_단원명", "")) or None,
            depth4_number=d4,
            depth4_name=str(row.get("Depth4_단원명", "")) or None,
            status=str(row.get("상태값", "")) or None,
            is_leaf=is_leaf,
        )
        db.add(lm)
        lm_count += 1

    await db.flush()

    # 학습맵 PK 조회
    lm_pk_map: dict[str, int] = {}
    result = await db.execute(select(LearningMap))
    for lm in result.scalars():
        lm_pk_map[lm.learning_map_id] = lm.id

    # N:M 매핑 생성
    for row in rows:
        lm_id = str(row.get("학습맵ID", "")).strip()
        std_ids_raw = str(row.get("표준체계 ID", "")).strip()

        if not lm_id or not std_ids_raw or lm_id not in lm_pk_map:
            continue

        lm_pk = lm_pk_map[lm_id]
        std_ids = [s.strip() for s in std_ids_raw.split("|") if s.strip()]

        for std_id in std_ids:
            if std_id not in std_id_map:
                continue

            std_pk = std_id_map[std_id]

            # 중복 체크
            existing = await db.execute(
                select(LearningMapStandard).where(
                    LearningMapStandard.learning_map_id == lm_pk,
                    LearningMapStandard.standard_id == std_pk,
                )
            )
            if existing.scalar_one_or_none():
                continue

            link = LearningMapStandard(
                learning_map_id=lm_pk,
                standard_id=std_pk,
            )
            db.add(link)
            link_count += 1

    logger.info(
        "learning_maps_imported",
        maps=lm_count,
        links=link_count,
    )


async def main():
    engine = create_async_engine(settings.DATABASE_URL, echo=False)
    async_session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with async_session() as db:
        print("📚 표준체계 임포트 중...")
        std_map = await import_curriculum_standards(db, CURRICULUM_XLSX)
        print(f"   ✅ {len(std_map)}개 표준체계 로드")

        print("🗺️  학습맵 임포트 중...")
        await import_learning_maps(db, LEARNING_MAP_XLSX, std_map)

        await db.commit()
        print("✅ 임포트 완료!")

    await engine.dispose()


if __name__ == "__main__":
    asyncio.run(main())
